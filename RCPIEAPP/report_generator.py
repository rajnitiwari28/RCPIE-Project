# report_generator.py
"""
Module for generating Excel reports for RC submissions
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from django.contrib.auth.models import User
from .models import (
    UserProfile, ResearchProposal, Patent, ProjectProposal, 
    CunsultancyProof, EnterprenuerProof, InnovationProof, ProposalProof
)
from datetime import datetime
from io import BytesIO


class RCReportGenerator:
    """Generate comprehensive Excel reports for RC submissions"""
    
    def __init__(self):
        self.wb = openpyxl.Workbook()
        self.wb.remove(self.wb.active)  # Remove default sheet
        self.style_header = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        self.style_header_font = Font(bold=True, color="FFFFFF", size=12)
        self.style_summary = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        self.style_summary_font = Font(bold=True, color="1F4E78", size=11)
        self.thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    def get_branches(self):
        """Get unique branches from UserProfile"""
        return UserProfile.objects.values_list('department', flat=True).distinct().order_by('department')
    
    def create_summary_sheet(self):
        """Create summary sheet with overall statistics"""
        ws = self.wb.create_sheet("Summary", 0)
        
        # Title
        ws['A1'] = "RC SUBMISSIONS REPORT - SUMMARY"
        ws['A1'].font = Font(bold=True, size=14, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color="203864", end_color="203864", fill_type="solid")
        ws.merge_cells('A1:F1')
        ws['A1'].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 25
        
        # Report Date
        ws['A3'] = f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws['A3'].font = Font(italic=True, size=10)
        
        # Branch-wise summary header
        row = 5
        headers = ['Branch/Department', 'Research Proposals', 'Patents', 'Projects', 'Consultancy', 'Entrepreneur', 'Innovation', 'Proposals', 'Total']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col)
            cell.value = header
            cell.font = self.style_header_font
            cell.fill = self.style_header
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = self.thin_border
        
        # Add data for each branch
        branches = self.get_branches()
        row = 6
        total_all = {'research': 0, 'patent': 0, 'project': 0, 'consultancy': 0, 'entrepreneur': 0, 'innovation': 0, 'proposal': 0}
        
        for branch in branches:
            research_count = ResearchProposal.objects.filter(user_profile__department=branch).count()
            patent_count = Patent.objects.filter(user_profile__department=branch).count()
            project_count = ProjectProposal.objects.filter(user_profile__department=branch).count()
            consultancy_count = CunsultancyProof.objects.filter(faculty__userprofile__department=branch).count()
            entrepreneur_count = EnterprenuerProof.objects.filter(faculty__userprofile__department=branch).count()
            innovation_count = InnovationProof.objects.filter(faculty__userprofile__department=branch).count()
            proposal_count = ProposalProof.objects.filter(faculty__userprofile__department=branch).count()
            
            total = research_count + patent_count + project_count + consultancy_count + entrepreneur_count + innovation_count + proposal_count
            
            # Update totals
            total_all['research'] += research_count
            total_all['patent'] += patent_count
            total_all['project'] += project_count
            total_all['consultancy'] += consultancy_count
            total_all['entrepreneur'] += entrepreneur_count
            total_all['innovation'] += innovation_count
            total_all['proposal'] += proposal_count
            
            values = [branch, research_count, patent_count, project_count, consultancy_count, entrepreneur_count, innovation_count, proposal_count, total]
            for col, value in enumerate(values, 1):
                cell = ws.cell(row=row, column=col)
                cell.value = value
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = self.thin_border
            row += 1
        
        # Totals row
        ws.cell(row=row, column=1).value = "TOTAL"
        ws.cell(row=row, column=1).font = self.style_summary_font
        ws.cell(row=row, column=1).fill = self.style_summary
        ws.cell(row=row, column=1).border = self.thin_border
        
        total_values = [
            total_all['research'], total_all['patent'], total_all['project'],
            total_all['consultancy'], total_all['entrepreneur'], total_all['innovation'],
            total_all['proposal'], sum(total_all.values())
        ]
        for col, value in enumerate(total_values, 2):
            cell = ws.cell(row=row, column=col)
            cell.value = value
            cell.font = self.style_summary_font
            cell.fill = self.style_summary
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = self.thin_border
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 25
        for col in range(2, 10):
            ws.column_dimensions[get_column_letter(col)].width = 15
    
    def create_branch_sheet(self, branch):
        """Create detailed sheet for each branch"""
        ws = self.wb.create_sheet(branch[:25])  # Excel sheet name limit
        
        # Title
        title_cell = ws['A1']
        title_cell.value = f"{branch} - DETAILED SUBMISSION REPORT"
        title_cell.font = Font(bold=True, size=13, color="FFFFFF")
        title_cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        ws.merge_cells('A1:H1')
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 22
        
        row = 3
        
        # Research Proposals
        research_data = self._get_research_proposals(branch)
        if research_data:
            row = self._add_section(ws, "RESEARCH PROPOSALS", research_data, row)
        
        # Patents
        patent_data = self._get_patents(branch)
        if patent_data:
            row = self._add_section(ws, "PATENTS", patent_data, row)
        
        # Project Proposals
        project_data = self._get_project_proposals(branch)
        if project_data:
            row = self._add_section(ws, "PROJECT PROPOSALS", project_data, row)
        
        # Consultancy
        consultancy_data = self._get_consultancy(branch)
        if consultancy_data:
            row = self._add_section(ws, "CONSULTANCY", consultancy_data, row)
        
        # Entrepreneur
        entrepreneur_data = self._get_entrepreneur(branch)
        if entrepreneur_data:
            row = self._add_section(ws, "ENTREPRENEUR", entrepreneur_data, row)
        
        # Innovation
        innovation_data = self._get_innovation(branch)
        if innovation_data:
            row = self._add_section(ws, "INNOVATION", innovation_data, row)
        
        # Proposals
        proposal_data = self._get_proposals(branch)
        if proposal_data:
            row = self._add_section(ws, "PROPOSALS", proposal_data, row)
        
        # Add statistics
        row += 2
        self._add_statistics(ws, branch, row)
        
        # Adjust column widths
        for col in range(1, 9):
            ws.column_dimensions[get_column_letter(col)].width = 18
    
    def _add_section(self, ws, section_title, data, start_row):
        """Add a section with data to the worksheet"""
        # Section title
        ws[f'A{start_row}'] = section_title
        ws[f'A{start_row}'].font = Font(bold=True, size=11, color="FFFFFF")
        ws[f'A{start_row}'].fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        ws.merge_cells(f'A{start_row}:H{start_row}')
        ws.row_dimensions[start_row].height = 18
        
        start_row += 1
        
        # Headers
        headers = ['RC_ID', 'Title', 'Faculty', 'Department', 'Status', 'Date', 'Email', 'Remarks']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=start_row, column=col)
            cell.value = header
            cell.font = self.style_header_font
            cell.fill = self.style_header
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = self.thin_border
        
        start_row += 1
        
        # Data rows
        for record in data:
            for col, value in enumerate(record, 1):
                cell = ws.cell(row=start_row, column=col)
                cell.value = value
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                cell.border = self.thin_border
            start_row += 1
        
        return start_row
    
    def _get_research_proposals(self, branch):
        """Get research proposals for a branch"""
        proposals = ResearchProposal.objects.filter(user_profile__department=branch)
        data = []
        for prop in proposals:
            data.append([
                prop.rc_id or 'N/A',
                prop.title,
                prop.user_profile.user.get_full_name() or prop.user_profile.user.username,
                branch,
                prop.status,
                prop.created_at.strftime('%Y-%m-%d') if prop.created_at else 'N/A',
                prop.user_profile.email,
                prop.co_author or ''
            ])
        return data
    
    def _get_patents(self, branch):
        """Get patents for a branch"""
        patents = Patent.objects.filter(user_profile__department=branch)
        data = []
        for patent in patents:
            data.append([
                patent.rc_id or 'N/A',
                patent.title,
                patent.user_profile.user.get_full_name() or patent.user_profile.user.username,
                branch,
                patent.status,
                patent.submission_date.strftime('%Y-%m-%d') if patent.submission_date else 'N/A',
                patent.email_id,
                patent.patent_type or ''
            ])
        return data
    
    def _get_project_proposals(self, branch):
        """Get project proposals for a branch"""
        projects = ProjectProposal.objects.filter(user_profile__department=branch)
        data = []
        for project in projects:
            data.append([
                project.rc_id or 'N/A',
                project.title,
                project.user_profile.user.get_full_name() or project.user_profile.user.username,
                branch,
                project.status,
                project.user_profile.user.date_joined.strftime('%Y-%m-%d') if project.user_profile else 'N/A',
                project.user_profile.email if project.user_profile else 'N/A',
                project.corresponding_author or ''
            ])
        return data
    
    def _get_consultancy(self, branch):
        """Get consultancy proofs for a branch"""
        consultancies = CunsultancyProof.objects.filter(faculty__userprofile__department=branch)
        data = []
        for consultancy in consultancies:
            faculty = consultancy.faculty
            user_profile = faculty.userprofile
            data.append([
                consultancy.rc_id or 'N/A',
                consultancy.cunsultancy_title,
                faculty.get_full_name() or faculty.username,
                branch,
                'Submitted',
                consultancy.uploaded_at.strftime('%Y-%m-%d') if consultancy.uploaded_at else 'N/A',
                faculty.email,
                'Consultancy'
            ])
        return data
    
    def _get_entrepreneur(self, branch):
        """Get entrepreneur proofs for a branch"""
        entrepreneurs = EnterprenuerProof.objects.filter(faculty__userprofile__department=branch)
        data = []
        for entrepreneur in entrepreneurs:
            faculty = entrepreneur.faculty
            data.append([
                entrepreneur.rc_id or 'N/A',
                entrepreneur.enterprenuer_title,
                faculty.get_full_name() or faculty.username,
                branch,
                'Submitted',
                entrepreneur.uploaded_at.strftime('%Y-%m-%d') if entrepreneur.uploaded_at else 'N/A',
                faculty.email,
                'Entrepreneur'
            ])
        return data
    
    def _get_innovation(self, branch):
        """Get innovation proofs for a branch"""
        innovations = InnovationProof.objects.filter(faculty__userprofile__department=branch)
        data = []
        for innovation in innovations:
            faculty = innovation.faculty
            data.append([
                innovation.rc_id or 'N/A',
                innovation.innovation_title,
                faculty.get_full_name() or faculty.username,
                branch,
                'Submitted',
                innovation.uploaded_at.strftime('%Y-%m-%d') if innovation.uploaded_at else 'N/A',
                faculty.email,
                'Innovation'
            ])
        return data
    
    def _get_proposals(self, branch):
        """Get proposal proofs for a branch"""
        proposals = ProposalProof.objects.filter(faculty__userprofile__department=branch)
        data = []
        for proposal in proposals:
            faculty = proposal.faculty
            data.append([
                proposal.rc_id or 'N/A',
                proposal.proposal_title,
                faculty.get_full_name() or faculty.username,
                branch,
                proposal.status,
                proposal.uploaded_at.strftime('%Y-%m-%d') if proposal.uploaded_at else 'N/A',
                faculty.email,
                'Proposal'
            ])
        return data
    
    def _add_statistics(self, ws, branch, start_row):
        """Add statistics for the branch"""
        # Title
        ws[f'A{start_row}'] = "BRANCH STATISTICS"
        ws[f'A{start_row}'].font = Font(bold=True, size=11, color="FFFFFF")
        ws[f'A{start_row}'].fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
        ws.merge_cells(f'A{start_row}:D{start_row}')
        ws.row_dimensions[start_row].height = 18
        
        start_row += 1
        
        # Faculty statistics
        ws[f'A{start_row}'] = "Faculty-wise Submissions:"
        ws[f'A{start_row}'].font = Font(bold=True, size=10)
        start_row += 1
        
        # Get unique faculties in this branch
        faculty_stats = {}
        
        research = ResearchProposal.objects.filter(user_profile__department=branch)
        for r in research:
            faculty_name = r.user_profile.user.get_full_name() or r.user_profile.user.username
            faculty_stats[faculty_name] = faculty_stats.get(faculty_name, 0) + 1
        
        patents = Patent.objects.filter(user_profile__department=branch)
        for p in patents:
            faculty_name = p.user_profile.user.get_full_name() or p.user_profile.user.username
            faculty_stats[faculty_name] = faculty_stats.get(faculty_name, 0) + 1
        
        for faculty_name, count in sorted(faculty_stats.items()):
            ws[f'A{start_row}'] = faculty_name
            ws[f'B{start_row}'] = count
            ws[f'A{start_row}'].border = self.thin_border
            ws[f'B{start_row}'].border = self.thin_border
            start_row += 1
        
        # Status summary
        start_row += 1
        ws[f'A{start_row}'] = "Status Summary:"
        ws[f'A{start_row}'].font = Font(bold=True, size=10)
        start_row += 1
        
        statuses = {}
        for prop in ResearchProposal.objects.filter(user_profile__department=branch):
            status = prop.status
            statuses[status] = statuses.get(status, 0) + 1
        
        for status, count in sorted(statuses.items()):
            ws[f'A{start_row}'] = status
            ws[f'B{start_row}'] = count
            ws[f'A{start_row}'].border = self.thin_border
            ws[f'B{start_row}'].border = self.thin_border
            start_row += 1
    
    def generate(self):
        """Generate the complete report"""
        # Create summary sheet
        self.create_summary_sheet()
        
        # Create sheets for each branch
        branches = self.get_branches()
        for branch in branches:
            self.create_branch_sheet(branch)
        
        # Return as bytes
        output = BytesIO()
        self.wb.save(output)
        output.seek(0)
        return output

