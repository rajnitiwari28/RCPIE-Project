import json
from django.contrib import messages  
from django.http import FileResponse, Http404, HttpResponseForbidden, JsonResponse
from django.shortcuts import get_object_or_404, render
# Create your views here.
from django.shortcuts import render, redirect
from django.contrib.auth import login, authenticate, logout
from django.urls import reverse

from RCPIEAPP.models import  ChatMessage, ResearchProposal, UserProfile
from .forms import ConsultancyProofForm, EnterprenuerProofForm, InnovationProofForm, UserCreationForm

from django.shortcuts import render, redirect
from django.contrib.auth import authenticate, login
from .forms import UserCreationForm
from .models import AcademicCoordinator, CunsultancyProof, DRC_Member, DRCMemberReview, DeanResearch, DepartmentDRC, EnterprenuerProof, InnovationProof, OtherDepartmentDRCHead, Patent_proof, ProposalProof, ProposalRouting, RCConvener, UserProfile
from RCPIEAPP import models


def register(request):
    if request.method == 'POST':
        form = UserCreationForm(request.POST)
        if form.is_valid():
            role = form.cleaned_data.get('role')

            # Create user
            user = form.save(commit=False)
            user.save()
            if role:
                group, created = Group.objects.get_or_create(name=role)
                user.groups.add(group)
  
            # ✅ Create UserProfile
            user_profile = UserProfile.objects.create(
                user=user,
                mob=form.cleaned_data.get('mob'),
                department=form.cleaned_data.get('department'),
                gender=form.cleaned_data.get('gender'),
                email=form.cleaned_data.get('email'),
                first_name=form.cleaned_data.get('first_name'),
                last_name=form.cleaned_data.get('last_name'),
                role=role,
                is_approved_by_admin=(role == 'Faculty')  # ✅ Faculty auto-approved only
            )

            # ✅ Create role-specific profile
            if role == 'Faculty':
                Faculty.objects.create(
                    user_profile=user_profile,
                    first_name=form.cleaned_data.get('first_name'),
                    last_name=form.cleaned_data.get('last_name'),
                    department=form.cleaned_data.get('department'),
                    email=form.cleaned_data.get('email'),
                    mob=form.cleaned_data.get('mob'),
                    gender=form.cleaned_data.get('gender')
                )

            elif role == 'RC Convener':
                # ✅ Simplified creation (no extra params)
                RCConvener.objects.create(user_profile=user_profile)

            elif role == 'Department DRC':
                DepartmentDRC.objects.create(user_profile=user_profile)

            elif role == 'Other Department DRC Head':
                OtherDepartmentDRCHead.objects.create(user_profile=user_profile)

            elif role == 'Academic Coordinator':
                AcademicCoordinator.objects.create(user_profile=user_profile)

            elif role == 'DRC_Member':
                DRC_Member.objects.create(user_profile=user_profile)

            # ✅ Auto login only for Faculty
            if role == 'Faculty':
                user = authenticate(
                    username=form.cleaned_data.get('username'),
                    password=form.cleaned_data.get('password1')
                )
                if user:
                    login(request, user)
                    messages.success(request, "Your account has been successfully created! You are now logged in.")
                    return redirect('faculty_home')

            # ✅ All other roles → Pending approval
            messages.info(request, "Your account is pending admin approval. You will be able to log in once approved.")
            return redirect('approval_pending')

    else:
        form = UserCreationForm()

    return render(request, 'register.html', {'form': form})

def approval_pending(request):
    return render(request, 'approval_pending.html')  # Ensure this template exists


from django.contrib.auth.models import Group
from django.contrib.auth import authenticate, login
from django.shortcuts import redirect, render
from .models import UserProfile  # Assuming you have a UserProfile model

def login_view(request):
    if request.method == 'POST':
        username = request.POST['username']
        password = request.POST['password']
        user = authenticate(request, username=username, password=password)
        
        if user is not None:
            login(request, user)
            user_profile = UserProfile.objects.get(user=user)
            role = user_profile.role  # Fetch the user's role
            
            # Ensure the user is added to the correct group
            if role == 'Faculty':
                faculty_group, created = Group.objects.get_or_create(name='Faculty')
                user.groups.add(faculty_group)
                return redirect('faculty_home')  # Redirect to Faculty Home

            elif role == 'RC Convener':
                rc_group, created = Group.objects.get_or_create(name='RC_Convener')
                user.groups.add(rc_group)
                return redirect('rc_convener_home')  # Redirect to RC Convener Home

            elif role == 'Department DRC':
                drc_group, created = Group.objects.get_or_create(name='Department_DRC')
                user.groups.add(drc_group)
                return redirect('dept_drc_home')  # Redirect to Department DRC Home

            elif role == 'Other Department DRC Head':
                other_drc_group, created = Group.objects.get_or_create(name='Other_Dept_DRC_Head')
                user.groups.add(other_drc_group)
                return redirect('other_dept_drc_home')  # Redirect to Other Department DRC Head Home
            
            elif role == 'Academic Coordinator':
                academic_coordinator_group, created = Group.objects.get_or_create(name='Academic Coordinator')
                user.groups.add(academic_coordinator_group)
                return redirect('Academic_Coordinator_home') 
            elif role == 'Research_Dean':
                Research_Dean_group, created = Group.objects.get_or_create(name='Research_Dean')
                user.groups.add(Research_Dean_group)
                return redirect('dean_research_dashboard')  
            elif role == 'DRC_Member':
                DRC_Member_group, created = Group.objects.get_or_create(name='DRC_Member')
                user.groups.add(DRC_Member_group)
                return redirect('DRC_Member_dashboard') 
        else:
            return render(request, 'login.html', {'error': 'Invalid username or password'})

    return render(request, 'login.html')

from django.conf import settings
from .models import AcademicCoordinator

def academic_coordinator_dashboard(request):
    coordinator = AcademicCoordinator.objects.get(user_profile=request.user.userprofile)
    return render(request, "academic_coordinator_home.html", {"coordinator": coordinator})

from django.http import FileResponse
import mimetypes

def download_excel_file(request):
    coordinator = AcademicCoordinator.objects.get(user_profile=request.user.userprofile)
    file_path = coordinator.excel_file.path

    if os.path.exists(file_path):
        file_name = os.path.basename(file_path)
        mime_type, _ = mimetypes.guess_type(file_path)
        response = FileResponse(open(file_path, 'rb'), content_type=mime_type)
        response['Content-Disposition'] = f'attachment; filename="{file_name}"'
        return response

def upload_excel(request):
    """ Upload and save the Excel file in the AcademicCoordinator model """
    coordinator = AcademicCoordinator.objects.get(user_profile=request.user.userprofile)
    
    if request.method == "POST" and request.FILES.get("excel_file"):
        excel_file = request.FILES["excel_file"]

        # Delete previous file if it exists
        if coordinator.excel_file:
            if os.path.exists(coordinator.excel_file.path):
                os.remove(coordinator.excel_file.path)

        # Save new file
        coordinator.excel_file = excel_file
        coordinator.save()
        return redirect("Academic_Coordinator_home")

    return render(request, "academic_coordinator_home.html", {"coordinator": coordinator})

def delete_excel(request):
    """ Delete the Excel file from the model and storage """
    coordinator = AcademicCoordinator.objects.get(user_profile=request.user.userprofile)

    if coordinator.excel_file:
        file_path = coordinator.excel_file.path
        coordinator.excel_file.delete()
        coordinator.excel_file = None
        coordinator.save()

        # Delete the actual file
        if os.path.exists(file_path):
            os.remove(file_path)

    return JsonResponse({"message": "Excel file deleted successfully"})


from django.contrib.auth.decorators import login_required
@login_required(login_url='/login/')
def faculty_home(request):
    if not request.user.groups.filter(name='Faculty').exists():
        return HttpResponseForbidden("You are not allowed to access this page.")
    return render(request, 'faculty_home.html')

from django.contrib.auth.decorators import login_required
from RCPIEAPP.models import RCConvener, Notification, UserProfile, ResearchProposal
@login_required(login_url='/login/')

def research_submission(request):
    if request.method == 'POST':
        title = request.POST.get('title')
        abstract = request.POST.get('abstract')
        keywords = request.POST.get('keywords')
        area = request.POST.get('area')
        full_paper = request.FILES.get('full_paper')
        plagiarism = request.FILES.get('plagiarism')
        research_type = request.POST.get('research_type')
        CMRIT_Citation = request.POST.get('CMRIT_Citation')
        research_domain = request.POST.get('research_domain')

        user_profile = UserProfile.objects.get(user=request.user)

        # ✅ Create or update research proposal
        research, created = ResearchProposal.objects.get_or_create(
            user_profile=user_profile,
            title=title,
            defaults={
                'abstract': abstract,
                'keywords': keywords,
                'area': area,
                'full_paper': full_paper,
                'plagiarism': plagiarism,
                'research_type': research_type,
                'CMRIT_Citation': CMRIT_Citation,
                'research_domain': research_domain
            }
        )

        if not created:
            research.abstract = abstract
            research.keywords = keywords
            research.area = area
            research.research_type = research_type
            research.CMRIT_Citation = CMRIT_Citation
            research.research_domain = research_domain
            if full_paper:
                research.full_paper = full_paper
            if plagiarism:
                research.plagiarism = plagiarism
            research.save()

        # ✅ Always send notification to the correct RC user (“vijiv”)
        try:
            rc_user = User.objects.get(username='rccommon')  # The only RC Convener
            Notification.objects.create(
                sender=request.user,
                receiver=rc_user,
                message=(
                    f"New Research Proposal titled '{title}' submitted by "
                    f"{request.user.get_full_name() or request.user.username} for RC ID."
                ),
                link="/rc_convener_home/"
            )
            print("✅ Notification created for:", rc_user.username)
        except User.DoesNotExist:
            print("⚠️ RC Convener user 'vijiv' not found. Notification not sent.")

        return JsonResponse({'status': 'success', 'message': 'Updated' if not created else 'Created'})

    return JsonResponse({'status': 'error'})


from django.contrib.auth.decorators import login_required
from django.core.mail import send_mail
from django.conf import settings
@login_required(login_url='/login/')
def rc_id_generation(request):
    user_profile = UserProfile.objects.get(user=request.user)
    research_proposals = ResearchProposal.objects.filter(user_profile=user_profile)
    project_proposals = ProjectProposal.objects.filter(user_profile=user_profile)
    patent_proposals = Patent.objects.filter(user_profile=user_profile)
    pload_proposals = ProposalProof.objects.filter(faculty=user_profile.user)
    consultancy_proposals = CunsultancyProof.objects.filter(faculty=user_profile.user)
    enterprenuer_proposals = EnterprenuerProof.objects.filter(faculty=user_profile.user)
    innovation_proposals = InnovationProof.objects.filter(faculty=user_profile.user)
    users = User.objects.exclude(id=request.user.id)  # All other users

    # Extract RC_IDs and other relevant details
    proposals_with_rc_ids = []
    for proposal in research_proposals:
        proposals_with_rc_ids.append({
            'title': proposal.title,
            'status': proposal.status,
            'rc_id': proposal.rc_id,
        })
     
    project_with_rc_ids = []
    for project in project_proposals:
            project_with_rc_ids.append({
                'title': project.title,
                'status': project.status,
                'rc_id': project.rc_id,
            })
    patent_with_rc_ids = []
    for patent in patent_proposals:
            patent_with_rc_ids.append({
                'title': patent.title,
                'status': patent.status,
                'rc_id': patent.rc_id,
            })
    pload_rc_ids = []
    for pload in pload_proposals:
            pload_rc_ids.append({
                'title': pload.proposal_title,
                'rc_id': pload.rc_id,
            })
    consultancy_rc_ids = []
    for cload in consultancy_proposals:
            consultancy_rc_ids.append({
                'title': cload.cunsultancy_title,
                'rc_id': cload.rc_id,
            })
    enterprenuer_rc_ids = []
    for enterprenuer in enterprenuer_proposals:
            enterprenuer_rc_ids.append({
                'title': enterprenuer.enterprenuer_title,
                'rc_id': enterprenuer.rc_id,
            })
    innovation_rc_ids = []
    for innovation in innovation_proposals:
            enterprenuer_rc_ids.append({
                'title': innovation.innovation_title,
                'rc_id': innovation.rc_id,
            })
            
    # Send the information to the rc_id.html template
    return render(request, 'rc_id.html', {'proposals_with_rc_ids': proposals_with_rc_ids, 'project_with_rc_ids': project_with_rc_ids, 'patents_with_rc_ids':patent_with_rc_ids,'pload_rc_ids':pload_rc_ids,'consultancy_rc_ids':consultancy_rc_ids,'enterprenuer_rc_ids':enterprenuer_rc_ids,'innovation_rc_ids':innovation_rc_ids,'users':users})


from django.contrib.auth.decorators import login_required
@login_required(login_url='/login/')

def rc_convener_home(request):
    if not request.user.groups.filter(name='RC Convener').exists():
        return HttpResponseForbidden("You are not allowed to access this page.")

    # Fetch unread notifications
    notifications = Notification.objects.filter(receiver=request.user, is_read=False)
    # Optionally, separate by type if you want
    research_notifications = notifications.filter(notification_type='research')
    patent_notifications = notifications.filter(notification_type='patent')
    project_notifications = notifications.filter(notification_type='project')
    enterprenuer_notifications = notifications.filter(notification_type='enterprenuer')
    consultancy_notifications = notifications.filter(notification_type='consultancy')
    innovation_notifications = notifications.filter(notification_type='innovation')
    proposal_notifications = notifications.filter(notification_type='proposal')
    users = User.objects.exclude(id=request.user.id)  # exclude yourself
    return render(request, 'rc_convener_home.html', {
        'notifications': notifications,
        'research_notifications': research_notifications,
        'patent_notifications': patent_notifications,
        'project_notifications': project_notifications,
        'enterprenuer_notifications': enterprenuer_notifications,
        'consultancy_notifications': consultancy_notifications,
        'innovation_notifications': innovation_notifications,
        'proposal_notifications': proposal_notifications,
        'users': users,
    })
# Add in notification “Mark as Read”
@login_required
def mark_notification_read(request, notification_id):
    note = get_object_or_404(Notification, id=notification_id, receiver=request.user)
    note.is_read = True
    note.save()
    return redirect(note.link or 'rc_convener_home')


# Loading Research work on RC page  from Faculty page

from django.core.paginator import Paginator

from django.core.paginator import Paginator
from django.http import JsonResponse
from django.urls import reverse
import json

from django.contrib.auth.decorators import login_required
@login_required(login_url='/login/')
def load_proposals(request):

    department = request.GET.get('department')

    proposals = ResearchProposal.objects.all()

    if department and department != "All":
        proposals = proposals.filter(user_profile__department=department)

    proposals = proposals.order_by('-id')

    paginator = Paginator(proposals, 5)
    page_number = int(request.GET.get('page', 1))
    page_obj = paginator.get_page(page_number)

    proposals_list = []

    for p in page_obj:
        proposals_list.append({
            'id': p.id,
            'title': p.title,
            'abstract': p.abstract,
            'status': p.status,
            'user_profile__department': p.user_profile.department,
            'full_paper_url': p.full_paper.url if p.full_paper else None,
            'plagiarism_url': p.plagiarism.url if p.plagiarism else None,
            'dcr_comment': p.dcr_comment,
            'keywords': p.keywords,          
            'research_type': p.research_type,      
            })

    return JsonResponse({
        'proposals': proposals_list,
        'has_next': page_obj.has_next()
    })
@login_required(login_url='/login/')   
def view_proposal_details(request):
    proposal_id = request.GET.get('id')
    proposal = get_object_or_404(ResearchProposal, id=proposal_id)
    return render(request, 'proposal_details.html', {'proposal': proposal})

def serve_pdf_patent(request, patent_id):
    patent = get_object_or_404(Patent, id=patent_id)
    
    if not patent.full_pdf:
        raise Http404("PDF not found")

    file_path = patent.full_pdf.path  # Get the absolute file path

    response = FileResponse(open(file_path, 'rb'), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{os.path.basename(file_path)}"'
    
    return response




from django.contrib.auth.decorators import login_required
@login_required(login_url='/login/')
def serve_pdf(request, proposal_id):
    proposal = get_object_or_404(ResearchProposal, id=proposal_id)

    # Ensure the proposal has a full paper file
    if not proposal.full_paper:
        raise Http404("Full paper not found.")

    # Serve the file using FileResponse
    response = FileResponse(proposal.full_paper.open('rb'), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{proposal.title}.pdf"'
    
    return response
def serve_plagiarism(request, proposal_id):
    proposal = get_object_or_404(ResearchProposal, id=proposal_id)

    if not proposal.plagiarism:
        raise Http404("Plagiarism file not found.")

    response = FileResponse(proposal.plagiarism.open('rb'), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="${proposal.title}_plagiarism.pdf"'
    return response

from django.contrib.auth.decorators import login_required
@login_required(login_url='/login/')
def load_proposal_details(request):
    proposals = ResearchProposal.objects.all()
    return render(request, 'proposal_details.html', {'proposals': proposals})

#==========chat message====================

from django.http import JsonResponse
from .models import ChatMessage
from django.contrib.auth.models import User

def send_message(request):
    if request.method == "POST":
        sender = request.user
        receiver_id = request.POST.get("receiver_id")
        message = request.POST.get("message")
        receiver = User.objects.get(id=receiver_id)
        ChatMessage.objects.create(sender=sender, receiver=receiver, message=message)
        return JsonResponse({"success": True})

from django.views.decorators.csrf import csrf_exempt
from django.db.models import Q
from django.contrib.auth.models import User
from django.http import JsonResponse

@csrf_exempt
def get_messages(request):
    if request.method == "POST":
        user1 = request.user
        user2_id = request.POST.get("user2_id")
        user2 = User.objects.get(id=user2_id)

        messages = ChatMessage.objects.filter(
            Q(sender=user1, receiver=user2) |
            Q(sender=user2, receiver=user1)
        ).values("sender__username", "message", "timestamp")

        return JsonResponse(list(messages), safe=False)

from django.db.models import Q
from django.utils import timezone
from django.contrib.auth.decorators import login_required
@login_required(login_url='/login/')
def chat_view(request, receiver_username):
    # Get the receiver user
    receiver = User.objects.get(username=receiver_username)
    
    # Get the messages between the logged-in user (sender) and the receiver
    messages = ChatMessage.objects.filter(
        sender=request.user, receiver=receiver
    ) | ChatMessage.objects.filter(
        sender=receiver, receiver=request.user
    )
    messages = messages.order_by('timestamp')
    
    if request.method == 'POST':
        # Get the message from the POST request
        message = request.POST.get('message')
        if message:
            # Save the message to the database
            ChatMessage.objects.create(
                sender=request.user,
                receiver=receiver,
                message=message,
                timestamp=timezone.now()
            )
            return JsonResponse({'status': 'Message Sent'})
    
    return render(request, 'chat/chat.html', {'receiver': receiver, 'messages': messages})

from django.views.decorators.csrf import csrf_exempt  # if not using csrf token in AJAX
from django.contrib.auth.decorators import login_required

@login_required(login_url='/login/')
def fetch_chat_messages(request):
    if request.method == 'GET':
        messages = ChatMessage.objects.all().order_by('timestamp')
        data = [
            {
                'user': message.user.username,
                'message': message.message,
                'timestamp': message.timestamp.strftime('%Y-%m-%d %H:%M:%S'),
            }
            for message in messages
        ]
        return JsonResponse({'messages': data})
    return JsonResponse({'error': 'Invalid request method'}, status=400)

from django.contrib.auth.decorators import login_required
@login_required(login_url='/login/')
def delete_all_messages(request):
    if request.method == 'POST':
        # Delete all chat messages from the database
        ChatMessage.objects.all().delete()
        return JsonResponse({'success': True})

    return JsonResponse({'success': False, 'error': 'Invalid request method.'})
from django.contrib.auth.decorators import login_required
@login_required(login_url='/login/')
def load_chat(request):
    chat_messages = ChatMessage.objects.all()
    return render(request, 'chat_content.html', {'chat_messages': chat_messages})
from django.contrib.auth.decorators import login_required
@login_required(login_url='/login/')
def send_chat_message(request):
    if request.method == 'POST':
        message = request.POST.get('message')
        user_profile = UserProfile.objects.get(user=request.user)
        chat_message = ChatMessage(
            user_profile=user_profile,
            message=message
        )
        chat_message.save()
        return JsonResponse({'status': 'success'})
    return JsonResponse({'status': 'error'})

from django.contrib.auth.decorators import login_required
@login_required(login_url='/login/')
def update_proposal_status(request):
    if request.method == 'POST':
        proposal_ids = request.POST.get('proposal_ids', '').split(',')
        status = request.POST.get('status', '')
        for proposal_id in proposal_ids:
            proposal = get_object_or_404(ResearchProposal, id=proposal_id)
            proposal.status = status
            proposal.save()

        return JsonResponse({'message': 'Proposal status updated successfully.'})
    return JsonResponse({'error': 'Invalid request method.'}, status=400)

from django.contrib.auth.decorators import login_required
@login_required(login_url='/login/')
def send_to_dept_drc(request):
    if request.method == 'POST':
        proposal_ids = request.POST.getlist('proposal_ids')
        patent_ids = request.POST.getlist('patent_ids')
        project_ids = request.POST.getlist('project_ids[]')
        proposal_p_ids = request.POST.getlist('proposal_p_ids[]')
        patent_data = []
        for patent_id in patent_ids:
            patent = get_object_or_404(Patent, id=patent_id)
            patent.status = 'Sent to Dept DRC'
            patent.sent_to_department = patent.user_profile.department   # 👈 IMPORTANT

            patent.save()
            patent_data.append({
                'id': patent.id,
                'title': patent.title,
                'status': patent.status,
            })

        proposals_data = []
        for proposal_id in proposal_ids:
            proposal = get_object_or_404(ResearchProposal, id=proposal_id)
            proposal.status = 'Sent to Dept DRC'
            proposal.sent_to_department = proposal.user_profile.department   # 👈 IMPORTANT
            proposal.save()

            proposals_data.append({
                'id': proposal.id,
                'title': proposal.title,
                'abstract': proposal.abstract,
                'keywords': proposal.keywords,
                'status': proposal.status
            })
        projects_data = []
        for project_id in project_ids:
            project = get_object_or_404(ProjectProposal, id=project_id)
            project.status = 'Sent to Dept DRC'
            project.save()
            projects_data.append({
                'id': project.id,
                'title': project.title,
                'abstract': project.abstract,
                'status': project.status
            })
            
        # ----- Proposal Proof (P-Load) -----
        proposal_p_data = []
        for proof_id in proposal_p_ids:
            proof = get_object_or_404(ProposalProof, id=proof_id)
            # ⚡ Note: ProposalProof has no `status` field → you may add one if needed
            proposal_p_data.append({
                'id': proof.id,
                'title': proof.proposal_title,
                'rc_id': proof.rc_id,
                'proof_pdf_1': proof.proof_pdf_1.url if proof.proof_pdf_1 else None,
                'proof_pdf_2': proof.proof_pdf_2.url if proof.proof_pdf_2 else None,
                'faculty': proof.faculty.username,
                'uploaded_at': proof.uploaded_at.strftime("%Y-%m-%d %H:%M:%S"),
            })
        return JsonResponse({'status': 'success', 'proposals': proposals_data, 'patent': patent_data, 'projects': projects_data, 'proposal_p': proposal_p_data,})

    return JsonResponse({'status': 'error', 'message': 'Invalid request method'})


def load_proposal_p_drc(request):
    proposal_p_list = ProposalProof.objects.all()  # or filter by RC convener
    data = []
    for proof in proposal_p_list:
        data.append({
            'id': proof.id,
            'title': proof.proposal_title,
            'rc_id': proof.rc_id,
            'faculty': proof.faculty.username,
            'uploaded_at': proof.uploaded_at.strftime("%Y-%m-%d %H:%M:%S"),
            'proof_pdf_1': proof.proof_pdf_1.url if proof.proof_pdf_1 else None,
            'proof_pdf_2': proof.proof_pdf_2.url if proof.proof_pdf_2 else None,
        })
      
    return JsonResponse({'proposal_p': data})

def load_patent_to_drc(request):

    user_dept = request.user.userprofile.department

    patents = Patent.objects.filter(
        status='Sent to Dept DRC',
        user_profile__department=user_dept
    )
    
    data = []

    for patent in patents:

        if patent.full_pdf:
            full_pdf_url = patent.full_pdf.url
        else:
            full_pdf_url = None

        data.append({
            'id': patent.id,
            'title': patent.title,
            'abstract': patent.abstract,
            'keywords': patent.keywords,
            'status': patent.status,
            'dcr_comment': patent.dcr_comment,
            'user_profile_id': patent.user_profile_id,
            'full_pdf_url': full_pdf_url,
            "faculty_id": patent.user_profile.user.id,
        })

    return JsonResponse({'patents': data})
# Generate RC_ID for a proposal
from django.utils.timezone import now
from django.contrib.auth.decorators import login_required
from .models import Faculty, Patent, ProjectProposal, ResearchProposal, ProposalProof, CunsultancyProof, EnterprenuerProof, InnovationProof

@login_required(login_url='/login/')
def generate_rc_id(request):
    if request.method != 'POST':
        return JsonResponse({'status': 'error', 'message': 'Invalid request method'})

    rc_ids = []

    def generate_id(queryset, instance, code, department):
        current_year = now().year
        mon = now().strftime("%b").upper()
        existing_count = queryset.filter(
            rc_id__startswith=f"{current_year}_{mon}_{department}"
        ).count() + 1
        return f"{current_year}_{mon}_{department}_{code}{existing_count:03d}"

    def get_department(user_profile):
        return Faculty.objects.filter(user_profile=user_profile).values_list('department', flat=True).first()

    # ==== Research Proposals ====
    for proposal_id in request.POST.getlist('proposal_ids[]'):
        try:
            proposal = get_object_or_404(ResearchProposal, id=proposal_id)
            if proposal.rc_id:
                continue
            dept = get_department(proposal.user_profile)
            rc_id = generate_id(ResearchProposal.objects, proposal, '', dept)
            proposal.rc_id = rc_id
            proposal.save()
            rc_ids.append(rc_id)
        except Exception:
            continue

    # ==== Project Proposals ====
    for project_id in request.POST.getlist('project_ids[]'):
        try:
            project = get_object_or_404(ProjectProposal, id=project_id)
            if project.rc_id:
                continue
            dept = get_department(project.user_profile)
            rc_id = generate_id(ProjectProposal.objects, project, 'MEP', dept)
            project.rc_id = rc_id
            project.save()
            rc_ids.append(rc_id)
        except Exception:
            continue

    # ==== Patents ====
    for patent_id in request.POST.getlist('patent_ids[]'):
        try:
            patent = get_object_or_404(Patent, id=patent_id)
            if patent.rc_id:
                continue
            dept = get_department(patent.user_profile)
            rc_id = generate_id(Patent.objects, patent, 'T', dept)
            patent.rc_id = rc_id
            patent.save()
           
            rc_ids.append(rc_id)
        except Exception:
            continue

    # ==== Proposal Proofs ====
    for proposal_p_id in request.POST.getlist('proposal_p_ids[]'):
        try:
            proposal_p = get_object_or_404(ProposalProof, id=proposal_p_id)
            if proposal_p.rc_id:
                continue
            dept = proposal_p.faculty.userprofile.department
            rc_id = generate_id(ProposalProof.objects, proposal_p, 'P', dept)
            proposal_p.rc_id = rc_id
            proposal_p.save()
            rc_ids.append(rc_id)
        except Exception:
            continue

    # ==== Consultancy Proofs ====
    for consultancy_id in request.POST.getlist('consultancy_ids[]'):
        try:
            consultancy = get_object_or_404(CunsultancyProof, id=consultancy_id)
            if consultancy.rc_id:
                continue
            dept = consultancy.faculty.userprofile.department
            rc_id = generate_id(CunsultancyProof.objects, consultancy, 'C', dept)
            consultancy.rc_id = rc_id
            consultancy.save()
            rc_ids.append(rc_id)
        except Exception:
            continue

    # ==== Entrepreneur Proofs ====
    for enterprenuer_id in request.POST.getlist('enterprenuer_ids[]'):
        try:
            enterprenuer = get_object_or_404(EnterprenuerProof, id=enterprenuer_id)
            if enterprenuer.rc_id:
                continue
            dept = enterprenuer.faculty.userprofile.department
            rc_id = generate_id(EnterprenuerProof.objects, enterprenuer, 'E', dept)
            enterprenuer.rc_id = rc_id
            enterprenuer.save()
            rc_ids.append(rc_id)
        except Exception:
            continue

    # ==== Innovation Proofs ====
    for innovation_id in request.POST.getlist('innovation_ids[]'):
        try:
            innovation = get_object_or_404(InnovationProof, id=innovation_id)
            if innovation.rc_id:
                continue
            dept = innovation.faculty.userprofile.department
            rc_id = generate_id(InnovationProof.objects, innovation, 'E', dept)
            innovation.rc_id = rc_id
            innovation.save()
            rc_ids.append(rc_id)
        except Exception:
            continue

    # Final logic: Check how many were selected vs generated
    total_selected = (
        len(request.POST.getlist('proposal_ids[]')) +
        len(request.POST.getlist('patent_ids[]')) +
        len(request.POST.getlist('project_ids[]')) +
        len(request.POST.getlist('proposal_p_ids[]')) +
        len(request.POST.getlist('consultancy_ids[]')) +
        len(request.POST.getlist('enterprenuer_ids[]')) +
        len(request.POST.getlist('innovation_ids[]'))
    )

    if len(rc_ids) == 0 and total_selected > 0:
        return JsonResponse({
            'status': 'error',
        })

    request.session['rc_ids'] = rc_ids
    request.session.modified = True

    return JsonResponse({'status': 'success', 'rc_ids': rc_ids})


#===================update rc id==========================
@csrf_exempt
def update_rc_id(request):
    if request.method == 'POST':
        updates = json.loads(request.POST.get('rc_id_updates', '[]'))
        print("Updates received:", updates)
        errors = []

        for item in updates:
            old_rc_id = item.get('old')
            new_rc_id = item.get('new')

            if old_rc_id == new_rc_id:
                 # RC_ID not edited but still send email
                for Model in [ResearchProposal, ProjectProposal, Patent, ProposalProof, CunsultancyProof, EnterprenuerProof, InnovationProof]:
                    try:
                        obj = Model.objects.get(rc_id=old_rc_id)

                        # ===== SEND EMAIL =====
                        try:
                            if hasattr(obj, 'user_profile'):
                                email = obj.user_profile.email
                                title = obj.title
                            else:
                                email = obj.faculty.email
                                title = getattr(obj, 'proposal_title',
                                    getattr(obj, 'cunsultancy_title',
                                    getattr(obj, 'enterprenuer_title',
                                    getattr(obj, 'innovation_title', 'Proposal'))))

                            send_rcid_email(email, title, old_rc_id)

                        except Exception as e:
                            print("Email error:", e)

                        break
                    except Model.DoesNotExist:
                        continue
                continue  # No change, skip

            # Uniqueness check excluding the current RC_ID
            exists = any([
                ResearchProposal.objects.filter(rc_id=new_rc_id).exclude(rc_id=old_rc_id).exists(),
                ProjectProposal.objects.filter(rc_id=new_rc_id).exclude(rc_id=old_rc_id).exists(),
                Patent.objects.filter(rc_id=new_rc_id).exclude(rc_id=old_rc_id).exists(),
                CunsultancyProof.objects.filter(rc_id=new_rc_id).exclude(rc_id=old_rc_id).exists(),
                ProposalProof.objects.filter(rc_id=new_rc_id).exclude(rc_id=old_rc_id).exists(),
                EnterprenuerProof.objects.filter(rc_id=new_rc_id).exclude(rc_id=old_rc_id).exists(),
                InnovationProof.objects.filter(rc_id=new_rc_id).exclude(rc_id=old_rc_id).exists(),
            ])

            if exists:
                errors.append(f"{new_rc_id} already exists.We will assign default Id now")
                continue

            # Update in the correct model
            updated = False
            for Model in [ResearchProposal, ProjectProposal, Patent, ProposalProof, CunsultancyProof, EnterprenuerProof, InnovationProof]:
                try:
                    obj = Model.objects.get(rc_id=old_rc_id)
                    obj.rc_id = new_rc_id
                    obj.save()

                    # ===== SEND EMAIL AFTER FINAL SAVE =====
                    try:
                        if hasattr(obj, 'user_profile'):  # Research / Project / Patent
                            email = obj.user_profile.email
                            title = obj.title
                        else:  # Proof models
                            email = obj.faculty.email
                            title = getattr(
                                obj,
                                'proposal_title',
                                getattr(
                                    obj,
                                    'cunsultancy_title',
                                    getattr(
                                        obj,
                                        'enterprenuer_title',
                                        getattr(obj, 'innovation_title', 'Proposal')
                                    )
                                )
                            )

                        send_rcid_email(email, title, new_rc_id)

                    except Exception as e:
                        print("Email error:", e)

                    updated = True
                    break

                except Model.DoesNotExist:
                    continue

            if not updated:
                errors.append(f"{old_rc_id} not found.")

        if errors:
            return JsonResponse({'status': 'error', 'message': 'Some updates failed: ' + ', '.join(errors)})
        else:
            return JsonResponse({'status': 'success', 'message': 'RC_IDs updated successfully'})
    
    return JsonResponse({'status': 'error', 'message': 'Invalid request'})
#================Email Of RC_ID============================
from django.core.mail import send_mail
from django.conf import settings

def send_rcid_email(email, title, rc_id):
    subject = "RC_ID Generated for Your Proposal"

    plain_message = f"""
Dear Faculty,

Your proposal titled "{title}" has been assigned an RC_ID.

{rc_id}

Regards,
Research Committee
"""

    html_message = f"""
<p>Dear Faculty,</p>

<p>Your proposal titled "<b>{title}</b>" has been assigned an <b>RC_ID</b>.</p>

<p><b style="font-size:18px; color:blue;">{rc_id}</b></p>

<p>You can now track the progress using this RC_ID.</p>

<p>Regards,<br>Research Committee</p>
"""

    send_mail(
        subject,
        plain_message,   # fallback (important)
        settings.EMAIL_HOST_USER,
        [email],
        html_message=html_message,
        fail_silently=False
    )
# =======DEpartment DRC Work================================
from django.contrib.auth.decorators import login_required
@login_required(login_url='/login/')
def dept_drc_home(request):
    proposals = ResearchProposal.objects.all()
    return render(request, 'dept_drc_home.html', {
        'proposals': proposals
    })


from django.contrib.auth.decorators import login_required
@login_required(login_url='/login/')
def load_proposals_drc(request):

    user_dept = request.user.userprofile.department   # Logged-in DRC Department

    proposals = ResearchProposal.objects.filter(
        status='Sent to Dept DRC',
        user_profile__department=user_dept           # FILTER BY DEPARTMENT
    ).values('id', 'title', 'abstract', 'keywords','user_profile__user_id', 'status', 'full_paper','plagiarism','dcr_comment','drc_member_status','drc_member_comments')

    data = []

    for proposal in proposals:
        if proposal['full_paper']:
            proposal['full_paper_url'] = request.build_absolute_uri(reverse('serve_pdf', args=[proposal['id']]))
        else:
            proposal['full_paper_url'] = None
        if proposal['plagiarism']:
            proposal['plagiarism_url'] = request.build_absolute_uri(
                reverse('serve_pdf', args=[proposal['id']])
            )
        else:
            proposal['plagiarism_url'] = None

        data.append(proposal)

    return JsonResponse({'proposals': data})

from django.contrib.auth.decorators import login_required
@login_required(login_url='/login/')
def update_proposal_status(request):
    if request.method == 'POST':
        proposal_ids = request.POST.getlist('proposal_ids[]')
        status = request.POST.get('status')
        
        for proposal_id in proposal_ids:
            proposal = get_object_or_404(ResearchProposal, id=proposal_id)
            proposal.status = status
            proposal.save()
        return JsonResponse({'status': 'success'})
    return JsonResponse({'status': 'error'})

from django.core.mail import send_mail
from django.contrib.auth.decorators import login_required
@login_required(login_url='/login/')
def update_proposal_status_drc(request):
    if request.method == 'POST':
        proposal_id = request.POST.get('proposal_id')
        proposal_type = request.POST.get('proposal_type')
        status = request.POST.get('status')
        comment = request.POST.get('comment', '')
        if proposal_type == 'research':
            proposal = get_object_or_404(ResearchProposal, id=proposal_id)
        elif proposal_type == 'patent':
            proposal = get_object_or_404(Patent, id=proposal_id)
        elif proposal_type == 'project':
            proposal = get_object_or_404(ProjectProposal, id=proposal_id)
        elif proposal_type == 'proposal_p':  # ⚡ Added for Proposal Proof
            proposal = get_object_or_404(ProposalProof, id=proposal_id)
        else:
            return JsonResponse({'status': 'error', 'message': 'Invalid proposal type.'})

        # ⚡ Make sure ProposalProof model has these fields
        proposal.status = status
        if hasattr(proposal, 'dcr_comment'):
            proposal.dcr_comment = comment
        else:
            # If ProposalProof has no dcr_comment, you may add a field in model
            pass

        proposal.save()
        return JsonResponse({'status': 'success'})

    return JsonResponse({'status': 'error', 'message': 'Invalid request method.'})
def update_patent_status_drc(request):
    if request.method == 'POST':
        patent_id = request.POST.get('patent_id')
        patent_type = request.POST.get('patent_type')
        status = request.POST.get('status')
        comment = request.POST.get('comment', '')
        if patent_type == 'patent':
            patent = get_object_or_404(Patent, id=patent_id)
        else:
            return JsonResponse({'status': 'error', 'message': 'Invalid proposal type.'})

        # ⚡ Make sure ProposalProof model has these fields
        patent.status = status
        if hasattr(patent, 'dcr_comment'):
            patent.dcr_comment = comment
        else:
            # If ProposalProof has no dcr_comment, you may add a field in model
            pass

        patent.save()
        return JsonResponse({'status': 'success'})

    return JsonResponse({'status': 'error', 'message': 'Invalid request method.'})



from django.contrib.auth.models import User
from django.contrib.auth.decorators import login_required
@login_required(login_url='/login/')


def send_message_to_faculty(request):

    if request.method == "POST":

        proposal_id = request.POST.get('proposal_id')
        faculty_id = request.POST.get('faculty_id')
        message = request.POST.get('message')

        proposal = get_object_or_404(ResearchProposal, id=proposal_id)
        faculty = get_object_or_404(User, id=faculty_id)

        drc_email = request.user.email
        faculty_email = faculty.email

        send_mail(
            subject=f"Message regarding your proposal: {proposal.title}",
            message=message,
            from_email=drc_email,
            recipient_list=[faculty.email],
            fail_silently=False
        )
        print("✅ Email function executed successfully")

        return JsonResponse({'status': 'success'})

    return JsonResponse({'status': 'error'})

def send_message_to_faculty_patent(request):
    
    if request.method == "POST":

        patent_id = request.POST.get('patent_id')
        faculty_id = request.POST.get('faculty_id')
        message = request.POST.get('message')
        print("patent id", patent_id)
        print("faculty id", faculty_id)
        patent = get_object_or_404(Patent, id=patent_id)
        faculty = get_object_or_404(User, id=faculty_id)
        drc_email = request.user.email
        faculty_email = faculty.email
        send_mail(
            subject=f"Message regarding your proposal: {patent.title}",
            message=message,
            from_email=drc_email,
            recipient_list=[faculty.email],
            fail_silently=False
        )

        return JsonResponse({'status': 'success'})

    return JsonResponse({'status': 'error'})
# Other dept DRC views================================================================
from django.contrib.auth.decorators import login_required
def other_dept_drc_home(request):
    return render(request, 'other_dept_drc_home.html')
@login_required
def get_other_dept_drcs(request):
    current_dept = request.user.userprofile.department
    # Filter based on linked user's department
    drc_users = OtherDepartmentDRCHead.objects.exclude(
        user_profile__department=current_dept
    )

    data = [
        {
            "id": drc.user_profile.user.id,
            "name": f"{drc.user_profile.user.first_name} {drc.user_profile.user.last_name}",
            "department": drc.user_profile.department   # Correct department source
        }
        for drc in drc_users
    ]

    return JsonResponse({"drc_list": data})


from django.contrib.auth.decorators import login_required
@login_required(login_url='/login/')
def load_proposals_odrc(request):
    proposals = ResearchProposal.objects.filter(status='Sent to Other Dept DRC').values('id', 'title', 'abstract', 'keywords', 'status', 'full_paper','dcr_comment','plagiarism')
    data = []
    for proposal in proposals:
        if proposal['full_paper']:
            proposal['full_paper_url'] = request.build_absolute_uri(reverse('serve_pdf', args=[proposal['id']]))
        else:
            proposal['full_paper_url'] = None
        if proposal['plagiarism']:
            proposal['plagiarism_url'] = request.build_absolute_uri(
                reverse('serve_pdf', args=[proposal['id']])
            )
        else:
            proposal['plagiarism_url'] = None

        data.append(proposal)

    return JsonResponse({'proposals': data})

from django.contrib.auth.decorators import login_required
@login_required(login_url='/login/')

def update_proposal_status_odrc(request):
    if request.method == 'POST':
        proposal_id = request.POST.get('proposal_id')
        status = request.POST.get('status')  # "Approved" or "Rejected"

        comment = request.POST.get('comment', '')
        proposal = get_object_or_404(ResearchProposal, id=proposal_id)

        # ✅ Save ODRC decision
        proposal.odeptstatus = status
        proposal.dcr_comment = comment

        # ✅ IMPORTANT: CHANGE MAIN STATUS
        if status == 'Approved':
            proposal.status = 'Approved by Other Dept DRC'
        elif status == 'Rejected':
            proposal.status = 'Rejected by Other Dept DRC'

        proposal.save()

        print('Updated proposal', proposal.id, proposal.status)

        return JsonResponse({'odeptstatus': 'success'})

    return JsonResponse({'odeptstatus': 'error'})

from django.contrib.auth.decorators import login_required
@login_required(login_url='/login/')
def send_to_other_dept_drc(request):
    if request.method == 'POST':
        proposal_ids = request.POST.getlist('proposal_ids[]')
        drc_ids = request.POST.getlist('drc_ids[]')  # <-- Changed
        if not drc_ids:
            return JsonResponse({'status': 'error', 'message': 'No DRC selected'}, status=400)

        proposals_data = []

        for proposal_id in proposal_ids:
            proposal = get_object_or_404(ResearchProposal, id=proposal_id)
            proposal.status = 'Sent to Other Dept DRC'
            proposal.save()

            # Store routing history (so multiple DRCs can review)
            for drc_id in drc_ids:
                drc = get_object_or_404(OtherDepartmentDRCHead, user_profile_id=drc_id)
                ProposalRouting.objects.get_or_create(
                    proposal=proposal,
                    drc=drc,
                    defaults={'status': 'Pending Review'}
                )
                proposals_data.append({
                    'proposal_id': proposal.id,
                    'title': proposal.title,
                    'drc': f"{drc.user_profile.first_name} {drc.user_profile.last_name}",
                    'department': drc.department,
                    'status': proposal.status,
                })
                
        return JsonResponse({
            'status': 'success',
            'message': 'Proposal sent to selected DRC(s)',
            'proposals': proposals_data
        })

    return JsonResponse({'status': 'error', 'message': 'Invalid request'}, status=400)


from django.contrib.auth.decorators import login_required
@login_required(login_url='/login/')
def logout_view(request):
    logout(request)
    return redirect('login')


#==============patent work================================
import os
from django.contrib.auth.decorators import login_required
@login_required
def patent_submission(request):
    if request.method == 'POST' and request.FILES.get('full_pdf'):  # Ensure file is uploaded
        try:
            user_profile = UserProfile.objects.get(user=request.user)

            # Retrieve all fields from the form
            title = request.POST.get('patent_title')
            corresponding_author = request.POST.get('corresponding_author')
            co_author = request.POST.get('co_author')
            address = request.POST.get('address')
            email_id = request.POST.get('Email_id')
            abstract = request.POST.get('patent_abstract')
            keywords = request.POST.get('keywords')
            related_work = request.POST.get('related_work')
            prior_art = request.POST.get('prior_art')
            novelty = request.POST.get('novelty')
            full_pdf = request.FILES['full_pdf']  # File upload
            patent_type = request.POST.get('patent_type')

            # Validate required fields
            if not title or not abstract or not corresponding_author:
                return JsonResponse({'status': 'error', 'message': 'Missing required fields'})

            # Save the data to the database
            patent = Patent.objects.create(
                user_profile=user_profile,
                title=title,
                corresponding_author=corresponding_author,
                co_author=co_author,
                address=address,
                email_id=email_id,
                abstract=abstract,
                keywords=keywords,
                related_work=related_work,
                prior_art=prior_art,
                novelty=novelty,
                full_pdf=full_pdf,  # File field
                patent_type=patent_type,
            )
            
           # ------------------ Notification Logic ------------------
            rc_conveners = User.objects.filter(userprofile__role='RC Convener')
            for rc in rc_conveners:
                Notification.objects.create(
                    sender=request.user,
                    receiver=rc,
                    message=f"New Patent submitted by {request.user.get_full_name()}: {title}",
                    link=f"/rc_convener/view_patent/{patent.id}/",
                    notification_type='patent'  # Important!
                )
            # --------------------------------------------------------

            return JsonResponse({'status': 'success', 'message': 'Patent submitted successfully'})

        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)})

    return JsonResponse({'status': 'error', 'message': 'Invalid request'})




from django.contrib.auth.decorators import login_required
@login_required(login_url='/login/')
def load_patents(request):

    department = request.GET.get('department')

    patents = Patent.objects.all().order_by('-id')

    # Apply department filter
    if department and department != "All":
        patents = patents.filter(
            user_profile__department=department
        )

    patents_data = []

    for patent in patents:
        full_pdf_url = None

        if patent.full_pdf:
            full_pdf_url = request.build_absolute_uri(
                f"/serve_pdf_patent/{patent.id}/"
            )

        patents_data.append({
            "id": patent.id,
            "title": patent.title,
            "abstract": patent.abstract,
            "related_work": patent.related_work,
            "dcr_comment": patent.dcr_comment,
            "status": patent.status,
            "full_pdf": full_pdf_url,
            "patent_type": patent.patent_type,
            "department": patent.user_profile.department,
            "first_name": patent.user_profile.first_name,
            "last_name": patent.user_profile.last_name,
            "dean_comment": patent.dean_comment,
        })
        
    return JsonResponse({"patent": patents_data})

from django.contrib.auth.decorators import login_required

@login_required  # Ensure user is logged in
def upload_patent_proof(request):
    user_profile = request.user.userprofile  

    if request.method == "POST":
        patent_id = request.POST.get("patent_id")
        proof_link = request.POST.get("proof_link")
        patent_type = request.POST.get("patent_type")  # value from dropdown

        if patent_id:
            try:
                patent_obj = Patent.objects.get(id=patent_id, user_profile=user_profile)

                # Update proof link and type
                patent_obj.proof_link = proof_link
                patent_obj.patent_type = patent_type
                patent_obj.save()

                messages.success(request, "Patent proof updated successfully!")
                return redirect("upload_patent_proof")

            except Patent.DoesNotExist:
                messages.error(request, "Invalid patent selected")
                return redirect("upload_patent_proof")
        else:
            messages.error(request, "Please select a patent")
            return redirect("upload_patent_proof")

    # Fetch patents for dropdown
    patents = Patent.objects.filter(user_profile=user_profile)

    # Get patent type choices dynamically
    patent_type_choices = Patent.PATENT_TYPE_CHOICES

    return render(request, "patent_proof.html", {
        "patents": patents,
        "patent_type_choices": patent_type_choices
    })




from django.contrib.auth.decorators import login_required
@login_required(login_url='/login/')
def load_patents_drc(request):    
    # Fetch patents sent to DRC
    patents = Patent.objects.filter(status='Sent to Dept DRC').values('id', 'title', 'abstract', 'related_work', 'status')
    patents = list(patents)
    return JsonResponse({'patents': patents})

from django.contrib.auth.decorators import login_required
@login_required(login_url='/login/')
def update_patent_status_drc(request):
    if request.method == 'POST':
        patent_id = request.POST.get('patent_id')
        status = request.POST.get('status')
        comment = request.POST.get('comment', '')
        patent = get_object_or_404(Patent, id=patent_id)

        patent.status = status
       
        patent.dcr_comment = comment
        patent.save()
        return JsonResponse({'status': 'success'})
    return JsonResponse({'status': 'error'})

#======================Project Workflow =================

from django.views.decorators.csrf import csrf_exempt
from django.contrib.auth.decorators import login_required
@login_required(login_url='/login/')
@csrf_exempt

def project_submission(request):
    if request.method == 'POST':
        title = request.POST.get('project_title')
        abstract = request.POST.get('project_abstract')
        faculty_included = request.POST.get('faculty_included')
        proposal_form = request.FILES.get('proposal_form')
        user_profile = UserProfile.objects.get(user=request.user)
        try:
            # Create a new ProjectProposal instance and save it
            project = ProjectProposal.objects.create(
                user_profile=user_profile,

                title=title,
                abstract=abstract,
                faculty_included=faculty_included,
                proposal_form=proposal_form,

            )
            
            # ------------------ Notification Logic ------------------
            rc_conveners = User.objects.filter(userprofile__role='RC Convener')
            for rc in rc_conveners:
                Notification.objects.create(
                    sender=request.user,
                    receiver=rc,
                    message=f"New Project submitted by {request.user.get_full_name()}: {title}",
                    link=f"/rc_convener/view_project/{project.id}/",
                    notification_type='project'  # Important!
                )
            # --------------------------------------------------------
            project.save()

            return JsonResponse({'status': 'success'})

        except Exception as e:
            # Handle exceptions and return an error response
            return JsonResponse({'status': 'error', 'message': str(e)})
    
    return JsonResponse({'status': 'error', 'message': 'Invalid request method'})

# Load projects data on RC page from Faculty
from django.contrib.auth.decorators import login_required
@login_required(login_url='/login/')
def load_projects(request):

    # ✅ Get department from AJAX
    department = request.GET.get('department')

    # ✅ Base Query
    projects = ProjectProposal.objects.select_related(
        'user_profile'
    ).all().order_by('-id')

    # ✅ Apply Department Filter
    if department:
        projects = projects.filter(
            user_profile__department=department
        )

    # ✅ Required fields
    projects = projects.values(
        'id',
        'title',
        'abstract',
        'status',
        'proposal_form',
        'dcr_comment',
        'user_profile__department',
        'user_profile__first_name',
        'user_profile__last_name'
    )

    data = []

    for project in projects:

        # ✅ PDF URL
        if project['proposal_form']:
            proposal_url = request.build_absolute_uri(
                reverse('serve_pdf_project', args=[project['id']])
            )
        else:
            proposal_url = None

        data.append({
            "id": project['id'],
            "title": project['title'],
            "abstract": project['abstract'],
            "status": project['status'],
            "dcr_comment": project['dcr_comment'],
            "department": project['user_profile__department'],
            "first_name": project['user_profile__first_name'],
            "last_name": project['user_profile__last_name'],
            "proposal_form_url": proposal_url
        })

    return JsonResponse({'projects': data})

from django.contrib.auth.decorators import login_required
@login_required(login_url='/login/')
def serve_pdf_project(request, project_id):
    project = get_object_or_404(ProjectProposal, id=project_id)
    # Ensure the proposal has a full paper file
    if not project.proposal_form:
        raise Http404("Full paper not found.")

    # Serve the file using FileResponse
    response = FileResponse(project.proposal_form.open('rb'), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{project.title}.pdf"'
    
    return response
from django.contrib.auth.decorators import login_required
@login_required(login_url='/login/')
def load_projects_drc(request):
    
    projects = ProjectProposal.objects.filter(status='Sent to Dept DRC').values('id', 'title', 'abstract', 'faculty_included', 'status', 'proposal_form','dcr_comment')
    data = []

    for project in projects:
        if project['proposal_form']:
            project['proposal_form_url'] = request.build_absolute_uri(reverse('serve_pdf_project', args=[project['id']]))
        else:
            project['proposal_form_url'] = None
        data.append(project)
    return JsonResponse({'project': data})


from django.shortcuts import render
from .models import ResearchProposal
from django.contrib.auth.decorators import login_required
@login_required(login_url='/login/')
def view_all_research_proposals(request):
    if not request.user.is_authenticated:
        return redirect('login')

    proposals = ResearchProposal.objects.all().order_by('-id') 
    patents = Patent.objects.all().order_by('-id') 
    projects = ProjectProposal.objects.all().order_by('-id') 
    consultancy_proofs = CunsultancyProof.objects.all().order_by('-id')    
    proposal_proofs = ProposalProof.objects.all().order_by('-id') 
    enterprenuer_proofs = EnterprenuerProof.objects.all().order_by('-id')
    # 🔥 Attach title to each proof
    for proof in proposal_proofs:
        match = proposals.filter(rc_id=proof.rc_id).first()
        proof.main_title = match.title if match else "N/A"

    return render(request, 'view_all_research_proposals.html', {
        'proposal_proofs': proposal_proofs,
        'proposals': proposals,
        'patents': patents,
        'projects': projects,
        'consultancy_proofs':consultancy_proofs,
        'enterprenuer_proofs':enterprenuer_proofs,
    })
#=================Published Paper====================

from .forms import PublishedPaperForm
from django.contrib.auth.decorators import login_required
@login_required(login_url='/login/')
# views.py
def upload_research_proof(request):
    if request.method == "POST":
        form = PublishedPaperForm(request.POST, user=request.user)

        if form.is_valid():
            paper = form.save(commit=False)
            paper.faculty = request.user   # set faculty manually
            paper.save()

            messages.success(request, "Form saved successfully!")
            return redirect('upload_research_proof')

        else:
            print(form.errors)

    else:
        form = PublishedPaperForm(user=request.user)

    return render(request, 'upload_research_proof.html', {'form': form})
from django.http import JsonResponse
from .models import ResearchProposal

def get_rc_id_research_proof(request):
    proposal_id = request.GET.get("proposal_id")

    try:
        proposal = ResearchProposal.objects.get(id=proposal_id)
        return JsonResponse({"rc_id": proposal.rc_id,"co_author":proposal.co_author})
    except ResearchProposal.DoesNotExist:
        return JsonResponse({"rc_id": "","co_author": ""})

from django.shortcuts import render, get_object_or_404
from django.http import Http404
from .models import PublishedPaper
from django.contrib.auth.decorators import login_required

@login_required(login_url='/login/')
def proof_details(request, rc_id):
    # Clean the rc_id (if necessary)
    
    rc_id = rc_id.strip()  # Remove leading/trailing spaces
    # Fetch all matching records
    published_papers = PublishedPaper.objects.values().filter(rcid=rc_id)
  
    published_papers=published_papers.first()

    return render(request, 'proof_details.html', {'published_papers': published_papers})

def proof_details_patent(request, rc_id):
 
    patents = Patent_proof.objects.filter(rc_id=rc_id)  
   
    return render(request, 'proof_details_patent.html', {'patents': patents, 'rc_id': rc_id})

#=====================project submissions =================
from django.shortcuts import render, redirect, get_object_or_404
from .models import ProjectProposal
from .forms import ProjectCompletionForm  # Ensure this form only has proof_link field
def upload_project_proof(request):
    if request.method == 'POST':
        form = ProjectCompletionForm(request.POST)
        if form.is_valid():
            rc_id = form.cleaned_data['rc_id']
            proof_link = form.cleaned_data['proof_link']

            # Check if the RC ID exists
            project = ProjectProposal.objects.filter(rc_id=rc_id).first()
            
            if project:
                project.proof_link = proof_link  # ✅ Store Google Drive link
                project.save()
                
                messages.success(request, 'Google Drive proof link saved successfully!')  
                return redirect('upload_project_proof')  # Redirect back to the form
            else:
                messages.error(request, 'RC ID not found. Please check and try again.')

    else:
        form = ProjectCompletionForm()

    return render(request, 'upload_project_completion.html', {'form': form})

def get_project_proof_link(request, rc_id):
    project = get_object_or_404(ProjectProposal, rc_id=rc_id)

    if project.proof_link:
        return redirect(project.proof_link)  # Redirect to the proof link (Google Drive, etc.)
    else:
        return HttpResponse("Proof link not available", status=404)

#=======Excel form================================

import openpyxl
from django.http import HttpResponse
from .models import ResearchProposal, Patent_proof, ProjectProposal  # Update with correct model names

def export_proposals_to_excel(request):
    # Create a new Excel workbook and add worksheets
    wb = openpyxl.Workbook()
    
    # Rename default sheet
    ws1 = wb.active
    ws1.title = "Research Proposals"

    # Create separate sheets for each proposal type
    ws2 = wb.create_sheet(title="Patent Proposals")
    ws3 = wb.create_sheet(title="Project Proposals")

    # Define column headers for Research Proposals
    ws1.append(["ID", "User Name", "Title", "RC_ID", "Status"])
    research_proposals = ResearchProposal.objects.all()
    for proposal in research_proposals:
        ws1.append([
            proposal.id,
            f"{proposal.user_profile.user.first_name} {proposal.user_profile.user.last_name}",
            proposal.title,
            proposal.rc_id,
            proposal.status
        ])

    # Define column headers for Patent Proposals
    ws2.append(["ID", "User Name", "Title", "Abstract", "RC_ID"])
    patents = Patent_proof.objects.all()
    for patent in patents:
        ws2.append([
            patent.id,
            f"{patent.user_profile.user.first_name} {patent.user_profile.user.last_name}",
            patent.title,
            patent.abstract,
            patent.rc_id
        ])

    # Define column headers for Project Proposals
    ws3.append(["ID", "User Name", "Faculty Involved", "Title", "Abstract", "RC_ID"])
    projects = ProjectProposal.objects.all()
    for project in projects:
        ws3.append([
            project.id,
            f"{project.user_profile.user.first_name} {project.user_profile.user.last_name}",
            project.faculty_included,
            project.title,
            project.abstract,
            project.rc_id
        ])

    # Create HTTP response with Excel file
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=Proposals_Data.xlsx'
    wb.save(response)
    return response

#============Proposal P==================

from .forms import ProposalProofForm
@login_required
def upload_proposal_proof(request):
    if request.method == 'POST':
        proposal_id = request.POST.get('proposal_id')  # from dropdown

        # ------------------ UPDATE EXISTING ------------------
        if proposal_id:
            try:
                proposal = ProposalProof.objects.get(id=proposal_id, faculty=request.user)

                # Only update proof_file
                if request.FILES.get('proof_file'):
                    proposal.proof_file = request.FILES.get('proof_file')

                proposal.save()

                messages.success(request, "Proof updated successfully!")
                return redirect('faculty_home')

            except ProposalProof.DoesNotExist:
                messages.error(request, "Invalid proposal selected")
                return redirect('faculty_home')

        # ------------------ CREATE NEW ------------------
        else:
            form = ProposalProofForm(request.POST, request.FILES)

            if form.is_valid():
                proposal_proof = form.save(commit=False)
                proposal_proof.faculty = request.user
                proposal_proof.save()

                # -------- Notification Logic --------
                rc_conveners = User.objects.filter(userprofile__role__iexact='RC Convener')
                for rc in rc_conveners:
                    Notification.objects.create(
                        sender=request.user,
                        receiver=rc,
                        message=f"New Proposal Proof uploaded by {request.user.get_full_name()}",
                        link=f"/rc_convener/view_proposal_proof/{proposal_proof.id}/",
                        notification_type='proposal'
                    )
                # -----------------------------------

                messages.success(request, "Proposal Proof uploaded successfully!")
                return redirect('faculty_home')

    else:
        form = ProposalProofForm()

    # 🔥 IMPORTANT: Send proposals for dropdown
    proposals = ProposalProof.objects.filter(
        faculty=request.user,
        rc_id__isnull=False   # Only those with RC_ID
    )

    return render(request, 'upload_proposal_proof.html', {
        'form': form,
        'proposals': proposals
    })

@login_required(login_url='/login/')

def load_proposal_p(request):

    department = request.GET.get('department')

    proposal_p = ProposalProof.objects.select_related(
        'faculty__userprofile'
    ).all().order_by('-id')

    if department:
        proposal_p = proposal_p.filter(
            faculty__userprofile__department=department
        )

    proposal_p_data = []

    for proposal in proposal_p:

        proof_pdf_1_url = None
        proof_pdf_2_url = None

        # ✅ Direct file URL (no view needed)
        if proposal.proof_pdf_1:
            proof_pdf_1_url = request.build_absolute_uri(proposal.proof_pdf_1.url)

        if proposal.proof_pdf_2:
            proof_pdf_2_url = request.build_absolute_uri(proposal.proof_pdf_2.url)

        faculty_user = proposal.faculty

        faculty_name = f"{faculty_user.first_name} {faculty_user.last_name}"

        faculty_department = (
            faculty_user.userprofile.department
            if hasattr(faculty_user, 'userprofile')
            else "Not Available"
        )

        proposal_p_data.append({
            "id": proposal.id,
            "proposal_title": proposal.proposal_title,
            "faculty_name": faculty_name,
            "department": faculty_department,
            "proof_pdf_1": proof_pdf_1_url,
            "proof_pdf_2": proof_pdf_2_url,
        })

    return JsonResponse({"proposal_p": proposal_p_data})

def serve_pdf_proposal_p(request, proposal_p_id):
    proposal_p = get_object_or_404(ProposalProof, id=proposal_p_id)
    if not proposal_p.proof_pdf_1 or not os.path.exists(proposal_p.proof_pdf_1.path):
        raise Http404("PDF not found")

    try:
        file_path = proposal_p.proof_pdf_1.path      
        response = FileResponse(open(file_path, 'rb'), content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{os.path.basename(file_path)}"'
        return response
    except Exception as e:
        raise Http404("Error retrieving PDF")
#=========================Consultancy========================
@login_required
def upload_consultancy_proof(request):
    if request.method == 'POST':
        consultancy_id = request.POST.get('consultancy_id')  # from dropdown

        # ------------------ UPDATE EXISTING ------------------
        if consultancy_id:
            try:
                obj = CunsultancyProof.objects.get(id=consultancy_id, faculty=request.user)

                if request.FILES.get('proof_pdf_1'):
                    obj.proof_pdf_1 = request.FILES.get('proof_pdf_1')

                if request.FILES.get('proof_file'):
                    obj.proof_file = request.FILES.get('proof_file')

                obj.save()

                messages.success(request, "Consultancy Proof updated successfully!")
                return redirect('faculty_home')

            except CunsultancyProof.DoesNotExist:
                messages.error(request, "Invalid Consultancy selected")
                return redirect('faculty_home')

        # ------------------ CREATE NEW ------------------
        else:
            form = ConsultancyProofForm(request.POST, request.FILES)

            if form.is_valid():
                obj = form.save(commit=False)
                obj.faculty = request.user
                obj.save()

                # -------- Notification --------
                rc_conveners = User.objects.filter(userprofile__role__iexact='RC Convener')
                for rc in rc_conveners:
                    Notification.objects.create(
                        sender=request.user,
                        receiver=rc,
                        message=f"Consultancy Proof uploaded by {request.user.get_full_name()}",
                        link=f"/rc_convener/view_consultancy_proof/{obj.id}/",
                        notification_type='consultancy'
                    )
                # -----------------------------

                messages.success(request, "Consultancy Proof uploaded successfully!")
                return redirect('faculty_home')

    else:
        form = ConsultancyProofForm()

    # 🔥 IMPORTANT: Send existing consultancy records
    consultancies = CunsultancyProof.objects.filter(faculty=request.user)

    return render(request, 'upload_consultancy_proof.html', {
        'form': form,
        'consultancies': consultancies
    })
@login_required(login_url='/login/')
def load_consultancy(request):

    # ✅ Get department from filter
    department = request.GET.get('department')

    # ✅ Optimized queryset
    consultancy = CunsultancyProof.objects.select_related(
        'faculty__userprofile'
    ).all().order_by('-id')

    # ✅ Apply department filter
    if department:
        consultancy = consultancy.filter(
            faculty__userprofile__department=department
        )

    consultancy_data = []

    for c in consultancy:

        proof_pdf_1_url = None

        if c.proof_pdf_1:
            proof_pdf_1_url = request.build_absolute_uri(
                f"/serve_pdf_consultancy/{c.id}/"
            )

        faculty_user = c.faculty

        faculty_name = f"{faculty_user.first_name} {faculty_user.last_name}"

        faculty_department = (
            faculty_user.userprofile.department
            if hasattr(faculty_user, 'userprofile')
            else "Not Available"
        )

        consultancy_data.append({
            "id": c.id,
            "consultancy_title": c.cunsultancy_title,
            "faculty_name": faculty_name,
            "department": faculty_department,
            "proof_pdf_1": proof_pdf_1_url,
        })

    return JsonResponse({"consultancy": consultancy_data})
def serve_pdf_consultancy(request, consultancy_id):
    consultancy = get_object_or_404(CunsultancyProof, id=consultancy_id)
    if not consultancy.proof_pdf_1 or not os.path.exists(consultancy.proof_pdf_1.path):
        raise Http404("PDF not found")

    try:
        file_path = consultancy.proof_pdf_1.path      
        response = FileResponse(open(file_path, 'rb'), content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{os.path.basename(file_path)}"'
        return response
    except Exception as e:
        raise Http404("Error retrieving PDF")

#=========================Enterprenuer========================
@login_required
def upload_enterprenuer_proof(request):
    if request.method == 'POST':
        entrepreneur_id = request.POST.get('entrepreneur_id')  # from dropdown

        # ------------------ UPDATE EXISTING ------------------
        if entrepreneur_id:
            try:
                obj = EnterprenuerProof.objects.get(
                    id=entrepreneur_id,
                    faculty=request.user
                )

                if request.FILES.get('proof_pdf_1'):
                    obj.proof_pdf_1 = request.FILES.get('proof_pdf_1')

                if request.FILES.get('proof_file'):
                    obj.proof_file = request.FILES.get('proof_file')

                obj.save()

                messages.success(request, "Entrepreneur Proof updated successfully!")
                return redirect('faculty_home')

            except EnterprenuerProof.DoesNotExist:
                messages.error(request, "Invalid Entrepreneur selected")
                return redirect('faculty_home')

        # ------------------ CREATE NEW ------------------
        else:
            form = EnterprenuerProofForm(request.POST, request.FILES)

            if form.is_valid():
                obj = form.save(commit=False)
                obj.faculty = request.user
                obj.save()

                # -------- Notification Logic --------
                rc_conveners = User.objects.filter(userprofile__role__iexact='RC Convener')
                for rc in rc_conveners:
                    Notification.objects.create(
                        sender=request.user,
                        receiver=rc,
                        message=f"Entrepreneur Proof uploaded by {request.user.get_full_name()}",
                        link=f"/rc_convener/view_enterprenuer_proof/{obj.id}/",
                        notification_type='other'
                    )
                # -----------------------------------

                messages.success(request, "Entrepreneur Proof uploaded successfully!")
                return redirect('faculty_home')

    else:
        form = EnterprenuerProofForm()

    # 🔥 IMPORTANT: send existing records to template
    entrepreneurs = EnterprenuerProof.objects.filter(faculty=request.user)

    return render(request, 'upload_enterprenuer_proof.html', {
        'form': form,
        'entrepreneurs': entrepreneurs
    })
@login_required(login_url='/login/')
def load_enterprenuer(request):

    # ✅ Get department filter
    department = request.GET.get('department')

    # ✅ Optimized queryset
    enterprenuer = EnterprenuerProof.objects.select_related(
        'faculty__userprofile'
    ).all().order_by('-id')

    # ✅ Apply department filter
    if department:
        enterprenuer = enterprenuer.filter(
            faculty__userprofile__department=department
        )

    enterprenuer_data = []

    for e in enterprenuer:

        proof_pdf_1_url = None

        if e.proof_pdf_1:
            proof_pdf_1_url = request.build_absolute_uri(
                f"/serve_pdf_enterprenuer/{e.id}/"
            )

        faculty_user = e.faculty

        faculty_name = f"{faculty_user.first_name} {faculty_user.last_name}"

        faculty_department = (
            faculty_user.userprofile.department
            if hasattr(faculty_user, 'userprofile')
            else "Not Available"
        )

        enterprenuer_data.append({
            "id": e.id,
            "enterprenuer_title": e.enterprenuer_title,
            "faculty_name": faculty_name,
            "department": faculty_department,
            "proof_pdf_1": proof_pdf_1_url,
        })

    return JsonResponse({"enterprenuer": enterprenuer_data})

def serve_pdf_enterprenuer(request, enterprenuer_id):
    enterprenuer = get_object_or_404(EnterprenuerProof, id=enterprenuer_id)
    if not enterprenuer.proof_pdf_1 or not os.path.exists(enterprenuer.proof_pdf_1.path):
        raise Http404("PDF not found")

    try:
        file_path = enterprenuer.proof_pdf_1.path      
        response = FileResponse(open(file_path, 'rb'), content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{os.path.basename(file_path)}"'
        return response
    except Exception as e:
        raise Http404("Error retrieving PDF")
    
#=========================Innovation========================
@login_required
def upload_innovation_proof(request):
    if request.method == 'POST':
        form = InnovationProofForm(request.POST, request.FILES)
        if form.is_valid():
            innovation_proof = form.save(commit=False)
            innovation_proof.faculty = request.user  # Associate faculty user
            innovation_proof.save()
              # ------------------ Notification Logic ------------------
            rc_conveners = User.objects.filter(userprofile__role__iexact='RC Convener')
            for rc in rc_conveners:
                Notification.objects.create(
                    sender=request.user,
                    receiver=rc,
                    message=f"New Innovation Proof uploaded by {request.user.get_full_name()}",
                    link=f"/rc_convener/view_innovation_proof/{innovation_proof.id}/",  # RC view URL
                    notification_type='innovation'  # You can create a new type like 'proof' if needed
                )
            # --------------------------------------------------------
            
            messages.success(request, "Innovation Proofs uploaded successfully!")
            return redirect('faculty_home')  # Redirect to faculty home page
    else:
        form = InnovationProofForm()
    
    return render(request, 'upload_innovation_proof.html', {'form': form})

@login_required(login_url='/login/')
def load_innovation(request):

    # ✅ Get department filter
    department = request.GET.get('department')

    # ✅ Optimized queryset
    innovation = InnovationProof.objects.select_related(
        'faculty__userprofile'
    ).all().order_by('-id')

    # ✅ Apply department filter
    if department:
        innovation = innovation.filter(
            faculty__userprofile__department=department
        )

    innovation_data = []

    for e in innovation:

        proof_pdf_1_url = None

        if e.proof_pdf_1:
            proof_pdf_1_url = request.build_absolute_uri(
                f"/serve_pdf_innovation/{e.id}/"
            )

        faculty_user = e.faculty

        faculty_name = f"{faculty_user.first_name} {faculty_user.last_name}"

        faculty_department = (
            faculty_user.userprofile.department
            if hasattr(faculty_user, 'userprofile')
            else "Not Available"
        )

        innovation_data.append({
            "id": e.id,
            "innovation_title": e.innovation_title,
            "faculty_name": faculty_name,
            "department": faculty_department,
            "proof_pdf_1": proof_pdf_1_url,
        })

    return JsonResponse({"innovation": innovation_data})
def serve_pdf_innovation(request, innovation_id):
    innovation = get_object_or_404(InnovationProof, id=innovation_id)
    if not innovation.proof_pdf_1 or not os.path.exists(innovation.proof_pdf_1.path):
        raise Http404("PDF not found")

    try:
        file_path = innovation.proof_pdf_1.path      
        response = FileResponse(open(file_path, 'rb'), content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{os.path.basename(file_path)}"'
        return response
    except Exception as e:
        raise Http404("Error retrieving PDF")
    
    
#=========================All Proposals data======================== 
from django.shortcuts import render
from .models import (
    ResearchProposal, Patent, ProjectProposal, Patent_proof,
    PublishedPaper, ProposalProof, CunsultancyProof, EnterprenuerProof,
    InnovationProof
)

import openpyxl
from django.http import HttpResponse
from .models import (
    PublishedPaper, ProposalProof, CunsultancyProof, 
    EnterprenuerProof, InnovationProof, Patent_proof
)

def export_all_data_excel(request):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # Remove default sheet

    # Map proof models to display names and relevant fields
    proof_models = {
        'Published Papers': {
            'model': PublishedPaper,
            'fields': ['title', 'faculty', 'rcid', 'co_authors', 'proof_link', 'published_date']
        },
        'Proposal Proofs': {
            'model': ProposalProof,
            'fields': ['proposal_title', 'faculty', 'rc_id', 'proof_pdf_1', 'proof_pdf_2', 'status']
        },
        'Consultancy Proofs': {
            'model': CunsultancyProof,
            'fields': ['cunsultancy_title', 'faculty', 'rc_id', 'proof_pdf_1', 'status']
        },
        'Entrepreneur Proofs': {
            'model': EnterprenuerProof,
            'fields': ['enterprenuer_title', 'faculty', 'rc_id', 'proof_pdf_1', 'status']
        },
        'Innovation Proofs': {
            'model': InnovationProof,
            'fields': ['innovation_title', 'faculty', 'rc_id', 'proof_pdf_1', 'status']
        },
        'Patent Proofs': {
            'model': Patent_proof,
            'fields': ['title', 'user_profile', 'rc_id', 'patent_number', 'proof_link', 'submitted_at']
        }
    }

    for sheet_name, data in proof_models.items():
        model = data['model']
        fields = data['fields']

        queryset = model.objects.all()

        # Group by department if possible
        department_groups = {}
        for obj in queryset:
            if hasattr(obj, 'user_profile'):
                dept = getattr(obj.user_profile, 'department', 'Unknown')
            elif hasattr(obj, 'faculty'):
                dept = getattr(obj.faculty.userprofile, 'department', 'Unknown')
            else:
                dept = 'General'

            if dept not in department_groups:
                department_groups[dept] = []
            department_groups[dept].append(obj)

        for dept, objs in department_groups.items():
            ws_title = f"{sheet_name[:15]}-{dept[:15]}"  # Excel sheet name limit
            ws = wb.create_sheet(title=ws_title)
            ws.append(fields)  # Header row

            for obj in objs:
                row = []
                for field in fields:
                    value = getattr(obj, field, '')
                    # If FileField or URLField, get URL
                    if hasattr(value, 'url'):
                        value = value.url
                    # If User or UserProfile, get username
                    elif hasattr(value, 'username'):
                        value = value.username
                    elif hasattr(value, 'user_profile'):
                        value = str(value.user_profile)
                    row.append(str(value))
                ws.append(row)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=proof_data.xlsx'
    wb.save(response)
    return response

    #====================Research dean=====================================
def dean_research_dashboard(request):
    dean = DeanResearch.objects.first()
    users = User.objects.exclude(id=request.user.id)  # exclude yourself
    return render(request, 'dean_research_dashboard.html', {'dean': dean, 'users':users})

def send_to_research_dean(request):
    if request.method == "POST":
        patent_ids = request.POST.getlist('patent_ids[]')
        for pid in patent_ids:
            patent = Patent.objects.get(id=pid)
            # Only allow I-Load Patent
            if patent.patent_type == "I load Patent":
                patent.assigned_to_research_dean = True
                patent.status = 'Sent to Research Dean'
                patent.save()
                
        return JsonResponse({"status": "success"})

    return JsonResponse({"status": "error"})

def load_research_dean_patents(request):
    patents = Patent.objects.filter(assigned_to_research_dean=True).order_by('-id')
    data = []
    for p in patents:
        data.append({
            "id": p.id,
            "title": p.title,
            "abstract": p.abstract,
            "keywords": p.keywords,
            "status": p.status,
            "patent_type": p.patent_type,
            "dean_comment": p.dean_comment,  
            "full_pdf": p.full_pdf.url if p.full_pdf else "",
        })
    return JsonResponse({"patents": data})
def update_research_dean_patent(request):
    if request.method == "POST":
        pid = request.POST.get("id")
        action = request.POST.get("action")
        dean_comment = request.POST.get("comment")        
        patent = Patent.objects.get(id=pid)
        if action == "approve":
            patent.status = "Approved by Research Dean"
        elif action == "reject":
            patent.status = "Rejected by Research Dean"
        patent.dean_comment = dean_comment
        patent.save()
        return JsonResponse({"status": "success"})


@login_required
def DRC_Member_dashboard(request):
    member = DRC_Member.objects.get(user_profile=request.user.userprofile)
    assigned_reviews = DRCMemberReview.objects.filter(drc_member=member).select_related('research_proposal')
    users = User.objects.exclude(id=request.user.id)  # exclude yourself

    return render(request, 'DRC_Member_dashboard.html', {'assigned_reviews': assigned_reviews,'users':users})


def get_drc_members(request):
    members = DRC_Member.objects.values('id', 'user_profile__user__first_name', 'department')
    member_list = [
        {'id': m['id'], 'name': m['user_profile__user__first_name'], 'department': m['department']}
        for m in members
    ]
    return JsonResponse({'members': member_list})

@csrf_exempt
def send_to_drc_member(request):
    if request.method == 'POST':
        proposal_ids = request.POST.getlist('proposal_ids[]')
        member_id = request.POST.get('member_id')

        member = DRC_Member.objects.get(id=member_id)
        users = User.objects.exclude(id=request.user.id)  # exclude yourself
        for pid in proposal_ids:
            proposal = ResearchProposal.objects.get(id=pid)
            # Create a DRCMemberReview entry (if not already)
            DRCMemberReview.objects.get_or_create(
                research_proposal=proposal,
                drc_member=member,
                defaults={'status': 'Pending'}
            )

        return JsonResponse({'status': 'success'})
    return JsonResponse({'status': 'error', 'message': 'Invalid request'}, status=400)


@csrf_exempt
@login_required
def submit_drc_review(request):
    if request.method == 'POST':
        review_id = request.POST.get('review_id')
        status = request.POST.get('status')
        comments = request.POST.get('comments')

        try:
            # Fetch the review object
            review = DRCMemberReview.objects.get(id=review_id)

            # Update Review Table
            review.status = status
            review.comments = comments
            review.reviewed_at = timezone.now()
            review.save()

            # Update related ResearchProposal
            proposal = review.research_proposal  # <-- get linked ResearchProposal
            proposal.drc_member_status = status
            proposal.drc_member_comments = comments
            proposal.save()

            return JsonResponse({
                'status': 'success',
                'message': 'Review submitted successfully and proposal updated!'
            })

        except DRCMemberReview.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': 'Review not found.'})
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)})


@login_required
def load_drc_members(request):
    members = DRC_Member.objects.select_related('user_profile__user').all()
    data = [
        {
            'id': m.id,
            'name': m.user_profile.user.get_full_name(),
            'department': m.department or '',
        }
        for m in members
    ]
    return JsonResponse({'members': data})


@login_required
def send_to_drc_members(request):
    if request.method == 'POST':
        members = json.loads(request.POST.get('members'))
        proposal_id = request.POST.get('proposal_id')
        proposal = ResearchProposal.objects.get(id=proposal_id)
        
        for member_id in members:
            member = DRC_Member.objects.get(id=member_id)
            DRCMemberReview.objects.create(
                research_proposal=proposal,
                drc_member=member,
                status='Pending'
            )

        return JsonResponse({'status': 'success', 'message': 'Proposal sent to members successfully!'})
